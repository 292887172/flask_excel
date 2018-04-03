from openpyxl import load_workbook

wb = load_workbook("ttt.xlsx")
# print(wb.sheetnames)    #['Sheet1', 'Sheet2']
sheet = wb['Sheet1']

'''

# 读取



print(sheet["A"]) # A 列对象
print(sheet["1"]) # 第一行对象

# (<Cell 'Sheet1'.A1>, <Cell 'Sheet1'.A2>, <Cell 'Sheet1'.A3>)
# (<Cell 'Sheet1'.A1>, <Cell 'Sheet1'.B1>, <Cell 'Sheet1'.C1>)


for i in sheet["A"]:
  print(i.value, end=" ")    # name syq ll      <-A列中的所有值

print(sheet["A1"].value)     # name ，第一行第一列对象



print(sheet.max_row)       # 3    <-最大行数
print(sheet.max_column)    # 3    <-最大列数

# 写入xlsx

'''

# 循环打印每行数据
for i in sheet.iter_rows():

  for j in i:
    print(j.value,end=" ")
  print()

