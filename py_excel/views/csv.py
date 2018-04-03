#!/usr/bin/env python
# -*- coding:utf-8 -*-
from flask import Blueprint,render_template,request,url_for
from openpyxl import load_workbook

csv = Blueprint('csv', __name__)

@csv.route('/csv',methods=['GET', "POST"],endpoint='file')
def csvview():

    if request.method == 'POST':
        try:
            file = request.files.get('execl_ttt')
            suffix = file.filename.split('.')

            # 如果这个文件存在并且是表格文件
            if file and suffix == 'xlsx' :

                f = r'C:\Users\sunyongqiang\PycharmProjects\flask_learn\py_excel\py_excel\media\%s' %file.filename
                file.save(f)
                wb = load_workbook(f)
                sheet = wb['Sheet1']
                for i in sheet.iter_rows():

                    for j in i:

                        print(j.value, end=" ")
                    print()

                return '上传成功'
            else:
                return '文件上传异常，请重新上传'
        except Exception as e:
            pass
    return render_template('csv.html')